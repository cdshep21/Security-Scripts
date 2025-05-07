#!venv\Scripts\python.exe
"""
Firewall Rule Parser and IP Verifier
------------------------------------
Automates the download, parsing, and highlighting of firewall rules that are a year or older.

Steps:
1. Retrieves credentials from 1Password
2. Downloads CSV files from firewall API
3. Parses valid IPs
4. Queries external API for verification timestamps
5. Outputs formatted Excel file

Dependencies: pandas, openpyxl, requests, 1Password CLI
Usage: python script.py --textfile firewall_list.txt
"""

import argparse
import requests
import os
import logging
import json
import csv
import re
import subprocess
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from typing import Tuple, List, Dict, Optional
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from contextlib import contextmanager
import configparser
from pathlib import Path

# Configuration
CONFIG = {
    'API_TIMEOUT': 30,
    'MAX_RETRIES': 3,
    'BACKOFF_FACTOR': 1,
    'BATCH_SIZE': 50,
    'FIREWALL_API_URL': 'https://firewall.example.com/api/fw_rules.php',
    'ASSETDB_API_URL': 'https://assetdb.example.com/api/arp_lookup.php',
    'VAULT_NAME': 'MyVaultName',
    'ITEM_TITLE': 'FirewallAPIUser'
}

def create_session() -> requests.Session:
    """Create a requests session with retry strategy."""
    session = requests.Session()
    retry_strategy = Retry(
        total=CONFIG['MAX_RETRIES'],
        backoff_factor=CONFIG['BACKOFF_FACTOR'],
        status_forcelist=[429, 500, 502, 503, 504]
    )
    adapter = HTTPAdapter(max_retries=retry_strategy)
    session.mount("http://", adapter)
    session.mount("https://", adapter)
    return session

@contextmanager
def safe_file_handle(filepath: str, mode: str = 'r'):
    """Context manager for safe file handling."""
    file_handle = None
    try:
        file_handle = open(filepath, mode)
        yield file_handle
    finally:
        if file_handle:
            file_handle.close()

def get_op(vault_name: str, item_title: str) -> Tuple[Optional[str], Optional[str]]:
    """
    Retrieve credentials from 1Password.
    
    Args:
        vault_name: Name of the 1Password vault
        item_title: Title of the item containing credentials
        
    Returns:
        Tuple of (username, password) or (None, None) if retrieval fails
    """
    try:
        result = subprocess.run(
            ["op", "item", "get", item_title, "--vault", vault_name, "--format", "json"],
            capture_output=True,
            text=True,
            check=True
        )
        item_json = json.loads(result.stdout)

        username = ""
        password = ""

        for field in item_json.get("fields", []):
            label = field.get("label", "").lower()
            if "username" in label:
                username = field.get("value")
            elif "password" in label:
                password = field.get("value")

        if not username or not password:
            raise ValueError("Missing username or password in 1Password item.")

        return username, password

    except subprocess.CalledProcessError as e:
        logging.error(f"Error retrieving 1Password item: {e.stderr}")
        return None, None
    except Exception as e:
        logging.error(f"Error parsing 1Password item: {e}")
        return None, None

def download_csv(firewall_name: str, username: str, password: str, firewall_folder: str, session: requests.Session) -> Optional[str]:
    """
    Download CSV file from firewall API.
    
    Args:
        firewall_name: Name of the firewall
        username: API username
        password: API password
        firewall_folder: Folder to save the CSV
        session: Requests session
        
    Returns:
        Path to downloaded CSV file or None if download fails
    """
    try:
        url = f"{CONFIG['FIREWALL_API_URL']}?fw={firewall_name}"
        response = session.get(
            url,
            auth=(username, password),
            timeout=CONFIG['API_TIMEOUT'],
            verify=True
        )
        response.raise_for_status()

        csv_filename = os.path.join(firewall_folder, f"{firewall_name}.csv")
        with safe_file_handle(csv_filename, 'wb') as f:
            f.write(response.content)
        logging.info(f"Downloaded CSV for firewall: {firewall_name}")
        return csv_filename
    except requests.exceptions.RequestException as e:
        logging.error(f"Failed to download CSV for firewall {firewall_name}: {e}")
        return None

def parse_csv(filename: str) -> Tuple[List[str], List[str], pd.DataFrame]:
    """
    Parse CSV file and extract valid IPs.
    
    Args:
        filename: Path to CSV file
        
    Returns:
        Tuple of (all IPs, unique IPs, DataFrame with valid rows)
    """
    try:
        with safe_file_handle(filename, 'r', encoding='utf-8') as csvfile:
            reader = csv.reader(csvfile)
            headers = next(reader)
            
            if not headers:
                raise ValueError("Empty CSV file")
                
            valid_columns = [i for i, header in enumerate(headers) if header and header.strip()][:7]
            valid_rows = []
            
            for row in reader:
                if len(row) >= max(valid_columns) + 1:
                    valid_row = [row[i] for i in valid_columns if i < len(row)]
                    valid_rows.append(valid_row)

        filtered_df = pd.DataFrame(valid_rows, columns=[headers[i] for i in valid_columns])

        if 'IP' not in filtered_df.columns:
            raise ValueError(f"No 'IP' column found in {filename}")

        filtered_df['IP'] = filtered_df['IP'].astype(str).str.strip()
        ip_pattern = re.compile(r'^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$')
        valid_ips_df = filtered_df[filtered_df['IP'].apply(lambda x: bool(ip_pattern.match(x)))]

        ip_list = valid_ips_df['IP'].tolist()
        unique_ip_list = list(set(ip_list))

        return ip_list, unique_ip_list, valid_ips_df
    except Exception as e:
        logging.error(f"Error parsing CSV file {filename}: {e}")
        return [], [], pd.DataFrame()

def query_ip(ip: str, session: requests.Session) -> str:
    """
    Query IP verification API.
    
    Args:
        ip: IP address to verify
        session: Requests session
        
    Returns:
        Verification timestamp or 'N/A' if query fails
    """
    try:
        url = f"{CONFIG['ASSETDB_API_URL']}?search_by=IP&search_term={ip}"
        response = session.get(url, timeout=CONFIG['API_TIMEOUT'], verify=True)
        response.raise_for_status()
        data = response.json()
        return data.get('verified', 'N/A')
    except Exception as e:
        logging.error(f"Error querying IP {ip}: {e}")
        return 'N/A'

def query_ips_batch(ip_batch: List[str], session: requests.Session) -> Dict[str, str]:
    """
    Query multiple IPs in a batch.
    
    Args:
        ip_batch: List of IPs to query
        session: Requests session
        
    Returns:
        Dictionary mapping IPs to verification timestamps
    """
    return {ip: query_ip(ip, session) for ip in ip_batch}

def query_ips_and_add_to_df(unique_ip_list: List[str], ip_list: List[str], 
                          valid_ips_df: pd.DataFrame, session: requests.Session) -> Tuple[pd.DataFrame, int, int]:
    """
    Query IPs and add verification data to DataFrame.
    
    Args:
        unique_ip_list: List of unique IPs
        ip_list: List of all IPs
        valid_ips_df: DataFrame with valid rows
        session: Requests session
        
    Returns:
        Tuple of (updated DataFrame, successful calls, failed calls)
    """
    verified_dict = {}
    successful_calls = 0
    failed_calls = 0
    rows_with_ips = valid_ips_df.copy()

    # Process IPs in batches
    for i in range(0, len(unique_ip_list), CONFIG['BATCH_SIZE']):
        batch = unique_ip_list[i:i + CONFIG['BATCH_SIZE']]
        batch_results = query_ips_batch(batch, session)
        verified_dict.update(batch_results)
        successful_calls += sum(1 for v in batch_results.values() if v != 'N/A')
        failed_calls += sum(1 for v in batch_results.values() if v == 'N/A')

    verified_list = [verified_dict[ip] for ip in ip_list]
    rows_with_ips['Verified'] = verified_list

    return rows_with_ips, successful_calls, failed_calls

def convert_and_format_excel(df, excel_filename):
    df.to_excel(excel_filename, index=False)
    workbook = load_workbook(excel_filename)
    sheet = workbook.active

    max_col = 8
    min_row = 1
    max_row = sheet.max_row

    while max_row > 1 and all(cell.value is None for cell in sheet[max_row]):
        max_row -= 1

    table_ref = f"A{min_row}:H{max_row}"
    tab = Table(displayName='FW_Rules', ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    sheet.add_table(tab)

    for col in range(1, max_col + 1):
        col_letter = sheet.cell(row=min_row, column=col).column_letter
        for row in range(min_row, max_row + 1):
            cell = sheet[f"{col_letter}{row}"]
            if col_letter not in ['A', 'H']:
                cell.alignment = Alignment(wrap_text=True)
            if col_letter in ['B', 'C', 'D', 'E', 'F', 'G']:
                cell.alignment = Alignment(vertical='top', horizontal='left', wrap_text=True)
            else:
                cell.alignment = Alignment(vertical='center', horizontal='center')

    sheet.row_dimensions[1].height = 20
    for i in range(2, max_row + 1):
        sheet.row_dimensions[i].height = 35

    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 30
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 60
    sheet.column_dimensions['F'].width = 20
    sheet.column_dimensions['G'].width = 20
    sheet.column_dimensions['H'].width = 20

    dull_green = "90EE90"
    dull_red = "FF6347"
    today = datetime.today()
    for row in sheet.iter_rows(min_row=2, max_row=max_row):
        cell = row[-1]
        if cell.value:
            try:
                cell_date = datetime.strptime(cell.value, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                continue
            days_diff = (today - cell_date).days
            fill_color = dull_green if days_diff <= 365 else dull_red
            for cell in row:
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    workbook.save(excel_filename)
    logging.info(f"Excel file saved and formatted as: {excel_filename}")

def export_parsed_data(firewall_name, ip_list, unique_ip_list, valid_ips_df, firewall_folder):
    results_folder = firewall_folder
    os.makedirs(results_folder, exist_ok=True)

    with open(os.path.join(results_folder, f"{firewall_name}_ip_list.json"), 'w') as f:
        json.dump(ip_list, f, indent=4)
    with open(os.path.join(results_folder, f"{firewall_name}_unique_ip_list.json"), 'w') as f:
        json.dump(unique_ip_list, f, indent=4)
    valid_ips_df.to_csv(os.path.join(results_folder, f"{firewall_name}_valid_rows.csv"), index=False)

def main():
    parser = argparse.ArgumentParser(description="Firewall rule parser")
    parser.add_argument('-t', '--textfile', help="Text file containing a list of firewall names")
    args = parser.parse_args()

    if not args.textfile:
        parser.error("Text file argument is required")

    text_file_path = args.textfile
    if not os.path.exists(text_file_path):
        parser.error(f"Text file not found: {text_file_path}")

    text_file_name = os.path.splitext(os.path.basename(text_file_path))[0]
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    top_level_folder = f"{text_file_name}_{timestamp}"
    os.makedirs(top_level_folder, exist_ok=True)

    central_log_file = os.path.join(top_level_folder, f"{text_file_name}.log")
    logging.basicConfig(
        filename=central_log_file,
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )

    try:
        with safe_file_handle(text_file_path, 'r') as f:
            firewall_names = [line.strip() for line in f if line.strip()]

        if not firewall_names:
            logging.error("No firewall names found in text file")
            return

        username, password = get_op(CONFIG['VAULT_NAME'], CONFIG['ITEM_TITLE'])
        if not username or not password:
            logging.error("Failed to retrieve credentials from 1Password.")
            return

        session = create_session()

        for firewall in firewall_names:
            firewall_folder = os.path.join(top_level_folder, firewall)
            os.makedirs(firewall_folder, exist_ok=True)

            firewall_log_file = os.path.join(firewall_folder, f"{firewall}.log")
            firewall_logger = logging.getLogger(firewall)
            handler = logging.FileHandler(firewall_log_file)
            handler.setLevel(logging.INFO)
            firewall_logger.addHandler(handler)

            try:
                csv_filename = download_csv(firewall, username, password, firewall_folder, session)
                if not csv_filename:
                    continue

                start_time = datetime.now()
                firewall_logger.info(f"Started processing {csv_filename} at {start_time}")

                ip_list, unique_ip_list, rows_with_ips = parse_csv(csv_filename)
                if not ip_list:
                    firewall_logger.info(f"No IPs were identified in {csv_filename}")
                    continue

                updated_rows_with_ips, successful_calls, failed_calls = query_ips_and_add_to_df(
                    unique_ip_list, ip_list, rows_with_ips, session
                )

                excel_filename = os.path.join(firewall_folder, f"{firewall}_rules.xlsx")
                convert_and_format_excel(updated_rows_with_ips, excel_filename)

                end_time = datetime.now()
                duration = end_time - start_time
                minutes, seconds = divmod(duration.total_seconds(), 60)
                
                firewall_logger.info(f"Completed processing {csv_filename} at {end_time}")
                firewall_logger.info(f"Duration: {int(minutes)} minutes {int(seconds)} seconds")
                firewall_logger.info(f"Total IPs Identified: {len(ip_list)}")
                firewall_logger.info(f"Successful API Calls: {successful_calls}")
                firewall_logger.info(f"Failed API Calls: {failed_calls}")
                firewall_logger.info(f"Output File: {excel_filename}")
                firewall_logger.info("=== End of Log ===")

            except Exception as e:
                firewall_logger.error(f"Error processing firewall {firewall}: {e}")
            finally:
                firewall_logger.removeHandler(handler)
                handler.close()

    except Exception as e:
        logging.error(f"An unexpected error occurred: {e}")
    finally:
        if 'session' in locals():
            session.close()

if __name__ == '__main__':
    main()
