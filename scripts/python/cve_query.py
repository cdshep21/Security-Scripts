# SPDX-License-Identifier: MIT
# Copyright (c) 2025 Corderius Shepherd

import os
import sys
import logging
import requests
import json
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from cryptography.fernet import Fernet
import urllib3
import re
from contextlib import contextmanager
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

# Uncomment to disable SSL warnings for self-signed certificates
# urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler("cve_query.log"),
        logging.StreamHandler()
    ]
)

# Configure requests session with retry strategy
def create_session():
    session = requests.Session()
    retry_strategy = Retry(
        total=3,
        backoff_factor=1,
        status_forcelist=[429, 500, 502, 503, 504]
    )
    adapter = HTTPAdapter(max_retries=retry_strategy)
    session.mount("http://", adapter)
    session.mount("https://", adapter)
    return session

@contextmanager
def safe_excel_workbook():
    wb = None
    try:
        wb = openpyxl.Workbook()
        yield wb
    finally:
        if wb:
            try:
                wb.close()
            except Exception as e:
                logging.error(f"Error closing workbook: {e}")

# Returns the directory where the script is located
def get_script_directory():
    return os.path.dirname(os.path.abspath(__file__))

# Loads the Fernet encryption key from an environment variable
def load_fernet_key():
    fernet_key = os.getenv("FERNET_KEY")
    if not fernet_key:
        raise ValueError("FERNET_KEY environment variable is not set.")
    return Fernet(fernet_key.encode())

# Loads the JSON config file containing Security Center info
def load_config_file(config_path):
    if not os.path.exists(config_path):
        raise FileNotFoundError(f"Config file not found: {config_path}")
    with open(config_path, "r") as config_file:
        return json.load(config_file)

# Decrypts an encrypted API key using the Fernet key
def decrypt_key(fernet, encrypted_key):
    try:
        return fernet.decrypt(encrypted_key.encode()).decode()
    except Exception as e:
        logging.error(f"Decryption failed: {e}")
        raise

# Queries Tenable.SC for plugin IDs associated with the specified CVE
def get_plugins_for_cve(tenable_url, headers, cert_path, cve, session):
    params = {"filterField": "xrefs", "op": "eq", "value": cve}
    try:
        response = session.get(f"{tenable_url}/rest/plugin", headers=headers, params=params, verify=cert_path)
        response.raise_for_status()
        plugins = response.json().get("response", {})
        if not plugins:
            logging.warning(f"No plugins found for CVE: {cve}")
            return []
        return [plugin["id"] for plugin in plugins if "id" in plugin]
    except requests.exceptions.RequestException as e:
        logging.error(f"Failed to retrieve plugin data for {cve}: {e}")
        return []

# Creates a query in Tenable.SC using the given plugin IDs and query type
def create_query(tenable_url, headers, cert_path, plugin_ids, cve, tool, session):
    payload = {
        'name': f'API Query for {cve} ({tool})',
        'description': f'This query was created via CVE_Query.py',
        'type': 'vuln',
        'tool': tool,
        'filters': [{'filterName': 'pluginID', 'operator': '=', 'value': ",".join(plugin_ids)}]
    }
    try:
        response = session.post(f"{tenable_url}/rest/query", headers=headers, json=payload, verify=cert_path)
        response.raise_for_status()
        return response.json()['response']['id']
    except requests.exceptions.RequestException as e:
        logging.error(f"Failed to create query ({tool}) for {cve}: {e}")
        return None

# Downloads analysis results from a query ID
def download_analysis(tenable_url, headers, query_id, cert_path, session):
    payload = {'type': 'vuln', 'query': {'id': query_id}, 'sourceType': 'cumulative'}
    try:
        response = session.post(f"{tenable_url}/rest/analysis", headers=headers, json=payload, verify=cert_path)
        response.raise_for_status()
        return response.json()['response']['results']
    except requests.exceptions.RequestException as e:
        logging.error(f"Failed to download analysis data: {e}")
        return []

# Adds a worksheet to the Excel file summarizing IP-level vulnerability data
def add_ip_summary_sheet(wb, name, data):
    sheet_name = f"{name}-IPSum"
    ws = wb.create_sheet(title=sheet_name)

    if not data:
        logging.warning(f"No data received for {sheet_name}.")
        return

    ws.append(["IP", "NetBIOS Name", "DNS Name", "Last Auth Run"])

    for item in data:
        excel_time = item["lastAuthRun"] / 86400 + 25569
        ws.append([item["ip"], item["netbiosName"], item["dnsName"], excel_time])

    for col in ws.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 5

    for row in ws.iter_rows(min_row=2, max_col=4, min_col=4):
        for cell in row:
            cell.number_format = "YYYY-MM-DD HH:MM:SS"
            cell.alignment = Alignment(horizontal="center", vertical="center")

# Adds a worksheet to the Excel file summarizing plugin-level vulnerability data
def add_vulnsum_sheet(wb, name, data):
    sheet_name = f"{name}-VulnSum"
    ws = wb.create_sheet(title=sheet_name)

    if not data:
        logging.warning(f"No data received for {sheet_name}.")
        return

    headers = ["Plugin ID", "Plugin Name", "Plugin Description", "Severity", "Repository", "IP Address", "DNS Name", "NetBIOS Name"]
    ws.append(headers)

    for entry in data:
        for host in entry.get("hosts", []):
            for ip_info in host.get("iplist", []):
                ws.append([
                    entry.get("pluginID", ""),
                    entry.get("name", ""),
                    entry.get("pluginDescription", ""),
                    entry.get("severity", {}).get("name", ""),
                    host.get("repository", {}).get("name", ""),
                    ip_info.get("ip", ""),
                    ip_info.get("dnsName", ""),
                    ip_info.get("netbiosName", "")
                ])

    table_range = f"A1:H{ws.max_row}"
    table = Table(displayName=f"VulnTable_{name}", ref=table_range)
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        width = max((len(str(cell.value)) for cell in col if cell.value), default=10) + 5
        ws.column_dimensions[col_letter].width = width if col_letter != "C" else 70
        for cell in col:
            align = Alignment(horizontal="left" if col_letter == "C" else "center", vertical="top" if col_letter == "C" else "center", wrapText=(col_letter == "C"))
            cell.alignment = align
            if cell.row == 1:
                cell.font = Font(bold=True)

    for row in ws.iter_rows(min_row=2):
        ws.row_dimensions[row[0].row].height = 50

# Generates a unique filename by appending a number if the target filename already exists
def get_unique_filename(base_path, base_name):
    name, ext = os.path.splitext(base_name)
    counter = 1
    while os.path.exists(os.path.join(base_path, base_name)):
        base_name = f"{name}_{counter}{ext}"
        counter += 1
    return base_name

# Main program execution
# Loads config, gets user input, queries Tenable.SC, and writes to Excel
def main():
    script_dir = get_script_directory()
    config_path = os.path.join(script_dir, "config.json")
    cert_path = os.path.join(script_dir, "dod_chain.crt")

    if not os.path.exists(config_path) or not os.path.exists(cert_path):
        logging.error("Required config or certificate file not found.")
        sys.exit(1)

    try:
        fernet = load_fernet_key()
        config = load_config_file(config_path)
        
        # Validate config structure
        required_keys = ['url', 'encrypted_key']
        for center_name, center_info in config.items():
            if not all(key in center_info for key in required_keys):
                raise ValueError(f"Invalid config structure for {center_name}")

        # Prompt user for CVE with format validation and retry logic
        MAX_ATTEMPTS = 3
        for attempt in range(MAX_ATTEMPTS):
            cve = input("Enter CVE (e.g., CVE-2023-1234): ").strip().upper()
            if not cve or cve.isspace():
                logging.warning("Empty CVE entered. Please try again.")
                continue
            if len(cve) > 20:  # Reasonable max length for CVE
                logging.warning("CVE too long. Please try again.")
                continue
            if re.match(r"^CVE-\d{4}-\d{4,}$", cve):
                logging.info(f"Valid CVE entered: {cve}")
                break
            else:
                logging.warning("Invalid CVE format. Expected format: CVE-YYYY-NNNN")
        else:
            logging.error("Maximum attempts exceeded. Exiting.")
            sys.exit(1)

        # Create session with retry strategy
        session = create_session()

        # Get plugin IDs from first configured Security Center
        first_center = next(iter(config.items()))
        tenable_url = first_center[1]['url']
        access_key, secret_key = decrypt_key(fernet, first_center[1]['encrypted_key']).split(";")
        headers = {'x-apikey': f'accesskey={access_key}; secretkey={secret_key}', 'Content-Type': 'application/json'}

        logging.info(f"Requesting plugin IDs for {cve} from {tenable_url}")
        plugin_ids = get_plugins_for_cve(tenable_url, headers, cert_path, cve, session)
        if not plugin_ids:
            sys.exit(1)

        with safe_excel_workbook() as wb:
            wb.remove(wb.active)

            # Iterate through each configured Security Center and generate reports
            for center_name, center_info in config.items():
                url = center_info['url']
                access_key, secret_key = decrypt_key(fernet, center_info['encrypted_key']).split(";")
                headers = {'x-apikey': f'accesskey={access_key}; secretkey={secret_key}', 'Content-Type': 'application/json'}

                logging.info(f"Creating IP summary query for {center_name}")
                query_id_ip = create_query(url, headers, cert_path, plugin_ids, cve, "sumip", session)
                data_ip = download_analysis(url, headers, query_id_ip, cert_path, session) if query_id_ip else []
                add_ip_summary_sheet(wb, center_name, data_ip)

                logging.info(f"Creating vulnerability summary query for {center_name}")
                query_id_vuln = create_query(url, headers, cert_path, plugin_ids, cve, "vulnipdetail", session)
                data_vuln = download_analysis(url, headers, query_id_vuln, cert_path, session) if query_id_vuln else []
                add_vulnsum_sheet(wb, center_name, data_vuln)

            # Save workbook to unique file
            output_file = get_unique_filename(script_dir, f"Query_Results_{cve}.xlsx")
            try:
                wb.save(os.path.join(script_dir, output_file))
                logging.info(f"Report saved as {output_file}")
            except PermissionError:
                logging.error(f"Cannot save file {output_file}. It may be open in another program.")
                sys.exit(1)

    except Exception as e:
        logging.error(f"An unexpected error occurred: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
